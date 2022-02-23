import telebot
from telebot import types
import json
import random,requests,datetime,sys,os,openpyxl
from openpyxl import load_workbook
from random import randint
from requests import request
from googletrans import Translator

token = "5130966692:AAFwk1ePFc1ua3MnOx80EE3MrhZ_xR6EUyM"
bot = telebot.TeleBot(token)

keyboard = types.ReplyKeyboardMarkup()
keyboard.row("!Расписание сегодня", '!Расписание завтра')
keyboard.add("/антиплагиат", "/translate","/погода")
keyboard.add("!Расписание больше")
keyboard2 = types.ReplyKeyboardMarkup()
keyboard2.row("!Расписание пн",'!Расписание вт', "!Расписание ср")
keyboard2.add('!Расписание чт', "!Расписание пт", '!Расписание сб')
keyboard2.add('!Расписание на неделю')

urllocal1=''

def weather1(message):
    if message.text.lower()=='на день':
        bot.send_message(message.chat.id, 'Погоду в каком городе вы хотите узнать')
        bot.register_next_step_handler(message, weather)
    elif message.text.lower()=='на неделю':
        bot.send_message(message.chat.id, 'Погоду в каком городе вы хотите узнать')
        bot.register_next_step_handler(message,weather2)
    else:
        send(message.chat.id,'Выбора нет ключ поверни и повтори с предоставленными вариантами')
def plagiat(message):
    try:
        msg = Translator().translate(message.text,dest='ru').text
        msg = Translator().translate(msg, dest='fi').text
        msg = Translator().translate(msg,dest='ru').text
        send(message.chat.id,msg)
    except BaseException:
        send(message.chat.id, 'Работает только с русским языком')
def weather(message):
    try:
        city=Translator().translate(message.text,dest='en').text
        appid='bbc65b998ba57073bdb3373908fc2df4'
        res = requests.get("http://api.openweathermap.org/data/2.5/weather",params={'q': city, 'units': 'metric', 'lang': 'ru', 'APPID': appid})
        data = res.json()
        weat=f'В городе: {message.text}.\nПогодные условия: {data.get("weather")[0].get("description")}.\nТемпература: {data.get("main").get("temp")}.\nМинимальная температура: {data.get("main").get("temp_min")}.\nМаксимальная температура: {data.get("main").get("temp_max")}.\n'
        send(message.chat.id,weat)
    except BaseException:
        send(message.chat.id,'К сожалению не удаётся найти погоду по вашему городу возможно вы его выдумали , но в любом случае сообщите об ошибке @enotoreket он обязательно всё проверит')

def weather2(message):
    try:
        city=Translator().translate(message.text,dest='en').text
        appid='bbc65b998ba57073bdb3373908fc2df4'
        res = requests.get("http://api.openweathermap.org/data/2.5/forecast",params={'q': city, 'units': 'metric', 'lang': 'ru', 'APPID': appid})
        data = res.json()
        for i in data.get('list'):
            weat=f'{i.get("dt_txt")}\nВ городе: {message.text}.\nПогодные условия: {i.get("weather")[0].get("description")}.\nТемпература: {i.get("main").get("temp")}.\nМинимальная температура: {i.get("main").get("temp_min")}.\nМаксимальная температура: {i.get("main").get("temp_max")}.\n'
            if i.get('dt_txt').find('12:00:00')!=-1:
                send(message.chat.id,weat)
    except BaseException as e:
        print(e)
        send(message.chat.id,'К сожалению не удаётся найти погоду по вашему городу возможно вы его выдумали , но в любом случае сообщите об ошибке @enotoreket он обязательно всё проверит')
def reg(message):
    bot.send_message(message.chat.id, 'Введите номер группы')
    global urllocal1
    urllocal1=message.text
    bot.register_next_step_handler(message, reg1)

def reg1(message):
    reg2(message.chat.id,message.text)

def reg2(ids,group):
    global urllocal1
    url=" "+urllocal1
    registr(ids,url,group)
def translate(message):
    if message.text.lower().find('ru')!=-1:
        bot.register_next_step_handler(message, translateru)
        send(message.chat.id, 'Напишите ваше предложение')
    elif message.text.lower().find('eng')!=-1:
        send(message.chat.id, 'Напишите ваше предложение')
        bot.register_next_step_handler(message, translateeng)
    else:
        send(message.chat.id, 'Бот не переводит ни на какие языки кроме русского и английского')
def translateeng(message):
    msg=Translator().translate(message.text,dest='en').text
    send(message.chat.id, msg)

def translateru(message):
    msg=Translator().translate(message.text,dest='ru').text
    send(message.chat.id, msg)

@bot.message_handler(commands=['start'])
def start(message):
    keyboard = types.ReplyKeyboardMarkup()
    keyboard.row("/time", "/help", '/url', "/time", '/me', "!расписание сегодня",'!расписание завтра',"!расписание пн",'!расписание вт',"!расписание ср",'!расписание чт',"!расписание пт",'!расписание сб',)
    bot.send_message(message.chat.id, 'Привет! Хочешь узнать Расписание?', reply_markup=keyboard)

@bot.message_handler(commands=['погода'])
def pogoda(message):
    keyboard3= types.ReplyKeyboardMarkup()
    keyboard3.row("на неделю", 'на день')
    bot.send_message(message.chat.id, 'На сколько дней?', reply_markup=keyboard3)
    bot.register_next_step_handler(message, weather1)


@bot.message_handler(commands=['reg'])
def registration(message):
    bot.send_message(message.chat.id, 'Введите ссылку на ваше рассписание')
    bot.register_next_step_handler(message,reg)

@bot.message_handler(commands=['антиплагиат'])
def antiplagiat(message):
    bot.send_message(message.chat.id, 'Введите ваш текст')
    bot.register_next_step_handler(message,plagiat)


@bot.message_handler(commands=['translate'])
def translatemain(message):
    keyboard3= types.ReplyKeyboardMarkup()
    keyboard3.row("ru", 'eng')
    bot.send_message(message.chat.id, 'На какой язык переводить ?', reply_markup=keyboard3)
    bot.register_next_step_handler(message,translate)

@bot.message_handler(commands=['time'])
def time(message):
    date=datetime.datetime.now()
    bot.send_message(message.chat.id, date, reply_markup=keyboard)

@bot.message_handler(commands=['url'])
def urlsendmtuci(message):
    ids=message.chat.id
    send(ids, 'Тогда тебе сюда – https://mtuci.ru/','chat')

@bot.message_handler(commands=['me'])
def start(message):
    ids=message.chat.id
    send(ids, f'Твой id {ids}')

@bot.message_handler(commands=['help'])
def help_message(message):
    ids=message.chat.id
    send(ids, 'Я умею скидывать расписание и текущее время также я могу отправить тебе ссылку на мтуси')


@bot.message_handler(content_types=['text'])
def answer(message):
    global keyboard2
    msg=message.text.lower()
    ids=message.chat.id
    if msg == "!время":
        send(ids, datetime.datetime.now(), 'chat')
    elif msg == "!обнови":
        update(ids)
    elif msg == "!перезапуск":
        send(ids, 'Перезапускаюсь,хозяин', 'chat')
        os.execl(sys.executable, 'python.exe', 'main.py', "привет")
    elif msg.find("!регистрация бота") != -1:
        x = msg.replace('!регистрация бота', '', 1)
        x = x.split('+')
        registr(ids, x[0], x[1])

    elif msg.find("!понедельник") != -1:
        opene(date(1), ids, 'chat')
    elif msg.find("!вторник") != -1:
        opene(date(2), ids, 'chat')
    elif msg.find("!среда") != -1:
        opene(date(3), ids, 'chat')
    elif msg.find("!четверг") != -1:
        opene(date(4), ids, 'chat')
    elif msg.find("!пятница") != -1:
        opene(date(5), ids, 'chat')
    elif msg.find("!суббота") != -1:
        opene(date(6), ids, 'chat')
    elif msg.find("!расписание на неделю") != -1:
        opene(date(0), ids, 'chat_all')
    elif msg.find('!расписание больше')!=-1:
        send(ids,"Какой день вас интересует",'chat',keyboard2)

    elif msg == "!расписание пн":
        opene(date(1), ids, 'chat')
    elif msg == "!расписание вт":
        opene(date(2), ids, 'chat')
    elif msg == "!расписание ср":
        opene(date(3), ids, 'chat')
    elif msg == "!расписание чт":
        opene(date(4), ids, 'chat')
    elif msg == "!расписание пт":
        opene(date(5), ids, 'chat')
    elif msg == "!расписание сб":
        opene(date(6), ids, 'chat')
    elif msg == "!расписание вс":
        opene(date(7), ids, 'chat')
    elif msg.find("расписание сегодня") != -1:
        opene(date(0), ids, 'chat')
    elif msg.find("!негр") != -1:
        send(ids, message, 'chat')
    elif msg.find("расписание завтра") != -1:
        opene(date(100), ids, 'chat')

    elif msg=='gym':
        for x in range(3):
            send(ids, 'GYM!', 'chat')
    elif msg == "бот":
        send(ids, 'Бот в здании <3\nС любовью кривой кодер Артём из Gym', 'chat')
    elif msg.find('gym?') != -1:
        send(ids,'GYM-это империя')
    elif (msg.find("мать") != -1) or (msg.find("мам") != -1):
        send(ids, "Про маму лишнее было")
    elif msg.find("ректор") != -1:
        send(ids, "слава МТУСИ")
    elif msg.find("дот") != -1:
        send(ids, "Бай май шеги ба ))")
    elif msg.find("спат") != -1:
        send(ids, "Спи спокойно сталкер")
    elif msg.find("жиза") != -1:
        send(ids, "Одобряю")
    elif msg.find("бин2105") != -1:
        send(ids, "Первый в мире полис GYM")
    elif msg.find("лес") != -1:
        send(ids, "Прекрати фармить лес, животное")
    elif msg.find("люблино") != -1:
        send(ids, "Метро люблино-работаем")
    elif msg == "!расписание вс":
        send(ids, "Какое расписание додик ты шо конч иди спатьки ")
    elif msg.find("жаль") != -1:
        send(ids, "Жаль, не взял с собой рундук, хехех, сундук для рун - РУНДУК")
    elif (msg == 'ты') or (msg.find(" ты ") != -1):
        send(ids, "Ты играешь по 20 мм в день. Зачем ? Обьяснишь мне ?")
    elif msg.find("интересно") != -1:
        send(ids, "Что тебе интересно а ?")

    elif msg.find("марг") != -1:
        send(ids,
                'Маргарита проявила невероятную поддержку и лояльность она вознаграждается почётным титулом подруга GYM')
    elif (msg.find("кат") != -1) or (msg.find("мельник") != -1):
        send(ids, 'Мельникова проявила невероятную тупость и идиотизм она позор БИН2104')


##ОТКРЫТИЕ ФАЙЛА
f = open('base.json','r+')
base = json.load(f)
##ПОЛУЧЕНИЕ НОВОЙ ССЫЛКИ
def geturl(url):
    t = request('GET', 'https://mtuci.ru/time-table').text
    r = open('mtuci.html', 'w')
    r.write(t)
    r.close()
    r= open('mtuci.html', 'r')
    urlend = ''
    count=0
    while urlend.lower().find(url.lower()) == -1:
        count+=1
        urlend = r.readline()
        if count>=2158:
            break

    try:
        urlend = urlend.split()
        urlend  = ''.join(urlend)
        urlend = urlend.replace('<li><h4>2.&nbsp;&nbsp;&nbsp;<ahref=\"', '', 1)
        urlend = urlend.replace('\"target=\"_blank\">', '', 1)
        urlend = f'https://mtuci.ru{urlend}'
        return urlend
    except BaseException:
        return False
##РЕГИСТРАЦИЯ ГРУППЫ
def registr(ids,url,groupnumber):
    url = url[35:]
    base[ids]={
        "url": url,
        "group": groupnumber
      }
    open('base.json','r+').close()
    f=open('base.json',"w")
    json.dump(base,f)
    f.close()
    updatereg(ids)
    send(ids,'Бот успешно зарегестрирован','chat')
    os.execl(sys.executable, 'python.exe', 'main.py', "привет")
    send(ids, 'Бот не был зарегестрирован пожалуйста обратитесь к @enotoreket', 'chat')
##ВЫЧИСЛЕНИЕ НУЖНОГО ДНЯ
def date(needday):
    a = datetime.datetime.now()+ datetime.timedelta(hours=4)
    if needday==100:
        a = a + datetime.timedelta(days=1)
    if (needday!=0)and(needday!=100):
        while a.timetuple().tm_wday!=needday-1:
            a = a + datetime.timedelta(days=1)
    day=a.timetuple().tm_wday
    yd=a.timetuple().tm_yday
    week=0
    if yd/7-yd//7>0.4:
        yd+=1
    yd=round(yd/7)
    if yd%2==1:
        week=1
    numbers=[day,week]
    return numbers
##ОТПРАВКА СООБЩЕНИЯ
def send(ids,text,ls="chat",keyboardd=keyboard):
    bot.send_message(ids,text,reply_markup=keyboardd)
##ОБНОВЛЕНИЕ РАСПИСАНИЯ В МОМЕНТ РЕГИСТРАЦИИ
def updatereg(ids):
    f = open('base.json', 'r+')
    base = json.load(f)
    f.close()
    url=base.get(str(ids)).get('url')
    url=geturl(url)
    if url!=False:
        r=requests.get(url,allow_redirects=True)
        try:
            open(f'{ids}.xlsx','wb').close
        except BaseException:
            pass
        open(f'{ids}.xlsx','wb').write(r.content)
##ОБНОВЛЕНИЕ РАСПИСАНИЯ
def update(ids):
    url=base.get(str(ids)).get('url')
    url=geturl(url)
    r=requests.get(url,allow_redirects=True)
    try:
        open(f'{ids}.xlsx','wb').close
    except BaseException:
        pass
    open(f'{ids}.xlsx','wb').write(r.content)
##РАСЧЁТ И ОТПРАВКА РАСПИСАНИЯ
def opene(dayw, ids, ls):
    day = dayw[0]
    week = dayw[1]
    member=ids
    groupnumber = int(base.get(str(ids)).get('group'))
    otvet = [' ', 'а я напомниаю что спонсор GYM →Parimatch', 'а GYM это империя',
             'можно послушать треки digital ocean', 'ы', 'гуляем', 'пьём пиво', 'идём в доту', 'идём в майн',
             'пьём пиво', 'можно поспать', 'можно пойти и выпить пиво', 'можно в майн',
             'можно в доту',  'пьём пиво', 'ПИВО', 'пьём пиво', 'можно поспать',
             'можно в майн', 'идём снимать кино', 'идём репетировать',
             'Gym это имеперия', ' ', ' ']
    try:
        wb = load_workbook(f'{ids}.xlsx')
        wb.active = groupnumber - 1
        sheet = wb.active
        g = ''
        alls = 0
        if ls.find('all') != -1:
            alls = 7
        if alls == 0:
            if day != 6:
                todayday = sheet.cell(row=13 + 6 * day, column=1).value
                g += todayday + '\n'
                for t in range(14 + 6 * day, 18 + 6 * day + 1):
                    timer = sheet.cell(row=t, column=3).value
                    c = sheet.cell(row=t, column=2).value
                    if week == 1:
                        columns = 7
                        praktcol = 5
                        prepodcol = 6
                    else:
                        columns = 8
                        praktcol = 10
                        prepodcol = 9
                    prepod = None
                    prakt = sheet.cell(row=t, column=praktcol).value
                    cell = sheet.cell(row=t, column=columns).value
                    urls = None
                    if cell != None:
                        if cell.find('Инженерная и компьютерная графика') != -1:
                            cell = cell.replace('Инженерная и компьютерная графика', 'Инж-Граф', 1)
                        if cell.find('Социология') != -1:
                            cell = cell.replace('Социология', 'Социология', 1)
                        if cell.find('Введение в информационные технологии') != -1:
                            cell = cell.replace('Введение в информационные технологии', 'Введение в ИТ', 1)
                        if cell.find('Физика') != -1:
                            cell = cell.replace('Физика', 'Физика', 1)
                        if cell.find('Высшая математика') != -1:
                            cell = cell.replace('Высшая математика', 'Вышмат', 1)
                        if cell.find('Теоретические основы электротехники') != -1:
                            cell = cell.replace('Теоретические основы электротехники', 'ТОЭ', 1)
                        if cell.find('Философия') != -1:
                            cell = cell.replace('Философия', 'Философия', 1)
                        if cell.find('Элективные дисциплины по физической культуре и спорту') != -1:
                            cell = cell.replace('Элективные дисциплины по физической культуре и спорту', 'Физ-ра', 1)
                        if cell.find('Иностранный язык') != -1:
                            cell = cell.replace('Иностранный язык', 'Ин яз', 1)
                            prepod = sheet.cell(row=t, column=prepodcol).value
                        if prepod != None:
                            o = f'[{c}.] ({timer}) ({prepod}) ({prakt}) {cell}'
                        else:
                            o = f'[{c}.] ({timer}) ({prakt}) {cell}'

                    else:
                        o = f'[{c}.] ({timer}) Пары нет {random.choice(otvet)}'
                    g = g + '\n' + o + '\n'
                    if urls != None:
                        g += f'{urls}\n'
            else:
                g = 'ВОСКРЕСЕНЬЕ\nПар нет Гуляем ))))'
            send(member,g,'chat')
        else:
            for day in range(alls):
                if day != 6:
                    todayday = sheet.cell(row=13 + 6 * day, column=1).value
                    g += todayday + '\n'
                    for t in range(14 + 6 * day, 18 + 6 * day + 1):
                        timer = sheet.cell(row=t, column=3).value
                        c = sheet.cell(row=t, column=2).value
                        if week == 1:
                            columns = 7
                            praktcol = 5
                            prepodcol = 6
                        else:
                            columns = 8
                            praktcol = 10
                            prepodcol = 9
                        prepod = None
                        prakt = sheet.cell(row=t, column=praktcol).value
                        cell = sheet.cell(row=t, column=columns).value
                        if cell != None:
                            if cell.find('Инженерная и компьютерная графика') != -1:
                                cell = cell.replace('Инженерная и компьютерная графика', 'Инж-Граф', 1)
                            if cell.find('Социология') != -1:
                                cell = cell.replace('Социология', 'Социология', 1)
                            if cell.find('Введение в информационные технологии') != -1:
                                cell = cell.replace('Введение в информационные технологии', 'Введение в ИТ', 1)
                            if cell.find('Физика') != -1:
                                cell = cell.replace('Физика', 'Физика', 1)
                            if cell.find('Высшая математика') != -1:
                                cell = cell.replace('Высшая математика', 'Вышмат', 1)
                            if cell.find('Теоретические основы электротехники') != -1:
                                cell = cell.replace('Теоретические основы электротехники', 'ТОЭ', 1)
                            if cell.find('Философия') != -1:
                                cell = cell.replace('Философия', 'Философия', 1)
                            if cell.find('Элективные дисциплины по физической культуре и спорту') != -1:
                                cell = cell.replace('Элективные дисциплины по физической культуре и спорту', 'Физ-ра',
                                                    1)
                            if cell.find('Иностранный язык') != -1:
                                cell = cell.replace('Иностранный язык', 'Ин яз', 1)
                                prepod = sheet.cell(row=t, column=prepodcol).value
                            if prepod != None:
                                o = f' &#12831{2 + c % 2}; [{c}.] ({timer}) ({prepod}) ({prakt}) {cell}'
                            else:
                                o = f' &#12831{2 + c % 2}; [{c}.] ({timer}) ({prakt}) {cell}'

                        else:
                            o = f' &#12831{2 + c % 2}; [{c}.] ({timer}) Пары нет {random.choice(otvet)}'
                        g = g + '\n' + o + '\n'
                else:
                    g = 'ВОСКРЕСЕНЬЕ\nПар нет Гуляем))))'
                send(member,g,'chat')
                g = ''

    except BaseException as err:
        print(err)
        send(member, f'Технические шоколадки обратитесь к создателю бота @enotoreket ошибка\n{err}',ls)
bot.polling(none_stop=True)